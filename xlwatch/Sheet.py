import openpyxl
import openpyxl.worksheet.worksheet
import xlwatch.Type

class Sheet:
    def set_table_with_key(self, table: dict, key: str, value: str):
        key_split = key.split('.')

        cur = table
        for name in key_split[:-1]:
            if not cur.get(name):
                cur[name] = {}
            cur = cur[name]
        cur[key_split[-1]] = value

    def set_type(self, key: str, value: str):
        self.set_table_with_key(self.type_table, key, value)

    def set_row(self, id: str, key: str, value: str):
        self.set_table_with_key(self.data_table[id], key, value)

    def stringify_as_lua_type(self, table, indent = 0):
        tmp = '{\n'
        for name in table:
            if type(table[name]) is dict:
                tmp += ('\t' * (indent + 1)) + f'{name}: {self.stringify_as_lua_type(table[name], indent + 1)},\n'
            else:
                tmp += ('\t' * (indent + 1)) + f'{name}: {table[name]},\n'
        
        if len(table) >= 1:
            tmp = tmp[:-2]
        tmp += '\n' + ('\t' * indent) + '}'

        return tmp

    def stringify_as_lua_table(self, table, indent = 0):
        tmp = '{\n'
        for name in table:
            if type(table[name]) is dict:
                tmp += ('\t' * (indent + 1)) + f'{name} = {self.stringify_as_lua_table(table[name], indent + 1)},\n'
            elif table[name] is None:
                tmp += ('\t' * (indent + 1)) + f'{name} = nil,\n'
            else:
                tmp += ('\t' * (indent + 1)) + f'{name} = {table[name]},\n'
        
        if len(table) >= 1:
            tmp = tmp[:-2]
        tmp += '\n' + ('\t' * indent) + '}'
        
        return tmp

    def stringify(self) -> str:
        tmp = 'export type Table_Type = '
        tmp += self.stringify_as_lua_type(self.type_table)
        tmp += '\n\nlocal Table : {['
        tmp += self.key_type
        tmp += ']: Table_Type} = '
        tmp += self.stringify_as_lua_table(self.data_table)
        tmp += '\n\nreturn Table'

        return tmp

    def __init__(self, name: str, sheet: openpyxl.worksheet.worksheet.Worksheet, types: dict[str, xlwatch.Type.Type]):
        self.name = name
        self.type_table = {}
        self.data_table = {}
        self.key_type = sheet.cell(2, 1).value
                
        for i in range(2, sheet.max_column + 1):
            if sheet.cell(1, i).value == None:
                continue
            
            _type = sheet.cell(2, i).value

            if _type in types:
                self.set_type(sheet.cell(1, i).value, types[_type].lua_type)

        for i in range(3, sheet.max_row + 1):
            id = sheet.cell(i, 1).value
            if id == None:
                continue
            if type(id) is int:
                id = f"[{id}]"
            
            self.data_table[id] = {}

            for j in range(2, sheet.max_column + 1):
                if sheet.cell(1, j).value == None:
                    continue

                _type = sheet.cell(2, j).value
                if _type in types:
                    if sheet.cell(i, j).value == None:
                        self.set_row(id, sheet.cell(1, j).value, 'nil')
                    else:
                        self.set_row(id, sheet.cell(1, j).value, types[_type].converter(sheet.cell(i, j)))
